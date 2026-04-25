#!/usr/bin/env python3
"""
usdm_deviation_report.py
────────────────────────
Compares a CDISC USDM 4.0 test-extraction JSON against a golden reference JSON
and produces a colour-coded Excel deviation report.

Usage
-----
  python usdm_deviation_report.py golden.json test.json [output.xlsx]

  If output path is omitted, the report is written next to the golden file as
  <golden_stem>_deviation_report.xlsx

Dependencies
------------
  pip install openpyxl

The script has no other third-party requirements.
"""

import sys
import json
import re
import argparse
import os
import shutil
import tempfile
import zipfile
from pathlib import Path
from collections import defaultdict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("ERROR: openpyxl is not installed. Run:  pip install openpyxl")


# ──────────────────────────────────────────────────────────────────────────────
# Colour palette
# ──────────────────────────────────────────────────────────────────────────────
C_DARK_BLUE   = "1F3864"
C_MED_BLUE    = "2E75B6"
C_LIGHT_BLUE  = "D6E4F0"
C_RED         = "C00000"
C_ORANGE      = "E26B0A"
C_AMBER       = "FFC000"
C_GREEN       = "375623"
C_LIGHT_GREEN = "E2EFDA"
C_LIGHT_RED   = "FCE4D6"
C_LIGHT_AMB   = "FFF2CC"
C_WHITE       = "FFFFFF"
C_GREY        = "F2F2F2"
C_DARK_GREY   = "595959"


# ──────────────────────────────────────────────────────────────────────────────
# Style helpers
# ──────────────────────────────────────────────────────────────────────────────
def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _border():
    s = Side(style='thin', color='BFBFBF')
    return Border(left=s, right=s, top=s, bottom=s)

def _center(wrap=False):
    return Alignment(horizontal='center', vertical='center', wrap_text=wrap)

def _left(wrap=True):
    return Alignment(horizontal='left', vertical='top', wrap_text=wrap)

def _pct_bg(v):
    """Background colour for a percentage value."""
    if v is None:  return C_LIGHT_BLUE
    if v >= 90:    return C_LIGHT_GREEN
    if v >= 50:    return C_LIGHT_AMB
    return C_LIGHT_RED

def _pct_fg(v):
    """Font colour for a percentage value."""
    if v is None:  return C_DARK_GREY
    if v >= 90:    return C_GREEN
    if v >= 50:    return "7F6000"
    return C_RED

def _cw(ws, col_index, width):
    ws.column_dimensions[get_column_letter(col_index)].width = width

def _hdr_row(ws, row, cols, bg=C_DARK_BLUE):
    for ci, val in enumerate(cols, 1):
        c = ws.cell(row=row, column=ci, value=val)
        c.font      = Font(name='Arial', bold=True, color=C_WHITE, size=10)
        c.fill      = _fill(bg)
        c.alignment = _center(wrap=True)
        c.border    = _border()
    ws.row_dimensions[row].height = 22

def _title_block(ws, row, text, ncols, sub=''):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name='Arial', bold=True, size=14, color=C_WHITE)
    c.fill = _fill(C_DARK_BLUE)
    c.alignment = _center()
    ws.row_dimensions[row].height = 28
    if sub:
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        c = ws.cell(row=row, column=1, value=sub)
        c.font = Font(name='Arial', size=10, color=C_WHITE)
        c.fill = _fill(C_MED_BLUE)
        c.alignment = _center()
        ws.row_dimensions[row].height = 16
    return row + 1

def _sect_hdr(ws, row, text, ncols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font      = Font(name='Arial', bold=True, size=10, color=C_WHITE)
    c.fill      = _fill(C_MED_BLUE)
    c.alignment = _left(wrap=False)
    c.border    = _border()
    ws.row_dimensions[row].height = 18
    return row + 1


# ──────────────────────────────────────────────────────────────────────────────
# USDM-specific analysis helpers
# ──────────────────────────────────────────────────────────────────────────────

# All USDM sections tracked for completeness (label, JSON key, level)
SD_SECTIONS = [
    ("Arms",                          "arms",                   "studyDesign"),
    ("Epochs",                        "epochs",                 "studyDesign"),
    ("Encounters",                    "encounters",             "studyDesign"),
    ("Activities",                    "activities",             "studyDesign"),
    ("Objectives",                    "objectives",             "studyDesign"),
    ("Estimands",                     "estimands",              "studyDesign"),
    ("Study Elements",                "elements",               "studyDesign"),
    ("Study Cells",                   "studyCells",             "studyDesign"),
    ("Eligibility Criteria (Design)", "eligibilityCriteria",    "studyDesign"),
    ("Analysis Populations",          "analysisPopulations",    "studyDesign"),
    ("Schedule Timelines",            "scheduleTimelines",      "studyDesign"),
    ("Biospecimen Retentions",        "biospecimenRetentions",  "studyDesign"),
]
VER_SECTIONS = [
    ("Study Identifiers",             "studyIdentifiers",       "studyVersion"),
    ("Study Titles",                  "titles",                 "studyVersion"),
    ("Elig. Criterion Items",         "eligibilityCriterionItems", "studyVersion"),
    ("Narrative Content Items",       "narrativeContentItems",  "studyVersion"),
    ("Abbreviations",                 "abbreviations",          "studyVersion"),
    ("Roles",                         "roles",                  "studyVersion"),
    ("Organizations",                 "organizations",          "studyVersion"),
    ("Study Interventions",           "studyInterventions",     "studyVersion"),
    ("Administ. Products",            "administrableProducts",  "studyVersion"),
    ("Biomedical Concepts",           "biomedicalConcepts",     "studyVersion"),
    ("BC Categories",                 "bcCategories",           "studyVersion"),
    ("BC Surrogates",                 "bcSurrogates",           "studyVersion"),
    ("Conditions",                    "conditions",             "studyVersion"),
]

REQUIRED_USDM_ENVELOPE_KEYS = {"study", "usdmVersion", "systemName", "systemVersion"}


def _cnt(obj, key):
    return len(obj.get(key, []))

def _pct_str(t, g):
    if g == 0: return "N/A" if t == 0 else "Extra"
    return f"{t/g*100:.0f}%"

def _pct_val(t, g):
    return None if g == 0 else round(t / g * 100, 1)


def _walk(obj, func):
    """Recursively walk every dict node in a JSON tree and call func(node)."""
    if isinstance(obj, dict):
        func(obj)
        for v in obj.values():
            _walk(v, func)
    elif isinstance(obj, list):
        for item in obj:
            _walk(item, func)


def instance_type_coverage(root):
    """Return (pct_with_instanceType, total_dicts, dicts_with_instanceType)."""
    total = has = 0
    def f(o):
        nonlocal total, has
        total += 1
        if "instanceType" in o:
            has += 1
    _walk(root, f)
    pct = round(has / total * 100, 1) if total else 0.0
    return pct, total, has


def id_format_stats(root):
    """
    Return (usdm_count, hash_count, total_ids).
    USDM sequential: TypeName_N  e.g. Objective_1
    Hash-based:      anything_XXXXXXXX (8 hex chars at end)
    """
    usdm = hash_ = total = 0
    def f(o):
        nonlocal usdm, hash_, total
        v = o.get("id")
        if isinstance(v, str):
            total += 1
            if re.match(r'^[A-Za-z]+_\d+$', v):          usdm  += 1
            elif re.match(r'^.+_[a-f0-9]{8}$', v):        hash_ += 1
    _walk(root, f)
    return usdm, hash_, total


def code_system_stats(root):
    """
    Return (cdisc_count, nci_count, other_count, total_codes).
    Only inspects objects where instanceType == 'Code'.
    """
    cdisc = nci = other = 0
    def f(o):
        nonlocal cdisc, nci, other
        if o.get("instanceType") == "Code" and o.get("codeSystem"):
            cs = o["codeSystem"]
            if "cdisc.org" in cs:                              cdisc += 1
            elif "NCI" in cs or "ncithesaurus" in cs.lower(): nci   += 1
            else:                                              other += 1
    _walk(root, f)
    total = cdisc + nci + other
    return cdisc, nci, other, total


def extension_attributes_coverage(root):
    """Return (pct_with_extensionAttributes, total_dicts, dicts_with_ea)."""
    total = has = 0
    def f(o):
        nonlocal total, has
        total += 1
        if "extensionAttributes" in o:
            has += 1
    _walk(root, f)
    pct = round(has / total * 100, 1) if total else 0.0
    return pct, total, has


def total_object_counts(g_sd, t_sd, g_ver, t_ver):
    g = t = 0
    for _, key, _ in SD_SECTIONS:  g += _cnt(g_sd, key); t += _cnt(t_sd, key)
    for _, key, _ in VER_SECTIONS: g += _cnt(g_ver, key); t += _cnt(t_ver, key)
    return g, t


def get_brief_title(ver):
    for title in ver.get("titles", []):
        if isinstance(title.get("type"), dict) and title["type"].get("code") == "C99905x1":
            return title.get("text", "")
    titles = ver.get("titles", [])
    return titles[0].get("text", "") if titles else None


def detect_study_key(root, prefer_lower=True):
    """Return 'study' or 'Study' depending on what's present in the root."""
    if "study" in root:  return "study"
    if "Study" in root:  return "Study"
    # Fallback: pick the key whose value contains a 'versions' list
    for k, v in root.items():
        if isinstance(v, dict) and "versions" in v:
            return k
    return None


def find_extra_toplevel_keys(root, study_key):
    """
    Return list of top-level keys that are NOT part of the USDM envelope.
    We treat the detected study key as equivalent to 'study' for this check.
    """
    allowed = REQUIRED_USDM_ENVELOPE_KEYS | {study_key}
    return sorted(k for k in root if k not in allowed)


# ──────────────────────────────────────────────────────────────────────────────
# Report builder
# ──────────────────────────────────────────────────────────────────────────────

def build_report(golden_path: Path, test_path: Path, output_path: Path):
    print(f"  Loading golden : {golden_path}")
    with open(golden_path, encoding="utf-8") as f:
        g_root = json.load(f)
    print(f"  Loading test   : {test_path}")
    with open(test_path, encoding="utf-8") as f:
        t_root = json.load(f)

    # ── Locate study objects ──────────────────────────────────────────────────
    g_key = detect_study_key(g_root)
    t_key = detect_study_key(t_root)
    if g_key is None:
        sys.exit("ERROR: Cannot find study object in golden JSON. "
                 "Expected a top-level 'study' or 'Study' key.")
    if t_key is None:
        sys.exit("ERROR: Cannot find study object in test JSON. "
                 "Expected a top-level 'study' or 'Study' key.")

    g_s   = g_root[g_key]
    t_s   = t_root[t_key]

    # Guard: at least one version and one studyDesign must exist
    if not g_s.get("versions"):
        sys.exit("ERROR: Golden JSON has no versions array on the study object.")

    g_ver = g_s["versions"][0]
    t_ver = t_s["versions"][0] if t_s.get("versions") else {}
    g_sd  = g_ver["studyDesigns"][0] if g_ver.get("studyDesigns") else {}
    t_sd  = (t_ver.get("studyDesigns") or [{}])[0]

    # ── Study label for sheet headers ─────────────────────────────────────────
    study_label = (
        g_s.get("name")
        or (g_ver.get("titles") or [{}])[0].get("text", "")
        or golden_path.stem
    )[:80]

    # ── Pre-compute all metrics ───────────────────────────────────────────────
    g_total, t_total = total_object_counts(g_sd, t_sd, g_ver, t_ver)
    overall_pct = round(t_total / g_total * 100, 1) if g_total else 0.0

    g_it_pct, g_it_tot, g_it_has = instance_type_coverage(g_s)
    t_it_pct, t_it_tot, t_it_has = instance_type_coverage(t_s)

    g_ea_pct, _, _ = extension_attributes_coverage(g_s)
    t_ea_pct, _, _ = extension_attributes_coverage(t_s)

    t_usdm_ids, t_hash_ids, t_tot_ids = id_format_stats(t_s)
    g_usdm_ids, g_hash_ids, g_tot_ids = id_format_stats(g_s)
    t_id_pct = round(t_usdm_ids / t_tot_ids * 100, 1) if t_tot_ids else 0.0
    g_id_pct = round(g_usdm_ids / g_tot_ids * 100, 1) if g_tot_ids else 0.0

    t_cdisc_c, t_nci_c, t_other_c, t_code_total = code_system_stats(t_s)
    g_cdisc_c, g_nci_c, g_other_c, g_code_total = code_system_stats(g_s)
    t_cdisc_pct = round(t_cdisc_c / t_code_total * 100, 1) if t_code_total else 0.0
    g_cdisc_pct = round(g_cdisc_c / g_code_total * 100, 1) if g_code_total else 0.0

    # Envelope compliance
    g_envelope_ok = all(k in g_root for k in REQUIRED_USDM_ENVELOPE_KEYS)
    t_envelope_ok = all(k in t_root for k in REQUIRED_USDM_ENVELOPE_KEYS)
    t_has_study_lower = "study" in t_root

    # Extra top-level keys in test
    t_extra_keys = find_extra_toplevel_keys(t_root, t_key)

    # Key sets for field diff
    g_top_keys = set(g_root.keys())
    t_top_keys = set(t_root.keys())
    g_ver_keys = set(g_ver.keys())
    t_ver_keys = set(t_ver.keys())
    g_sd_keys  = set(g_sd.keys())
    t_sd_keys  = set(t_sd.keys())

    # ── Build workbook ────────────────────────────────────────────────────────
    wb = Workbook()
    wb.remove(wb.active)

    # ═════════════════════════════════════════════════════════════════════════
    # SHEET 1  Executive Summary
    # ═════════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet("Executive Summary")
    ws.sheet_view.showGridLines = False
    for ci, w in enumerate([2, 28, 26, 36, 14, 14, 14, 14], 1): _cw(ws, ci, w)

    row = _title_block(ws, 1, "CDISC USDM 4.0 — Protocol Extraction Deviation Report",
                       8, sub=f"{study_label}")
    row += 1

    # Score card
    ws.merge_cells(f'B{row}:H{row}')
    ws.cell(row=row, column=2,
            value="OVERALL CONTENT COMPLETENESS SCORE").font = Font(
        name='Arial', bold=True, size=11, color=C_DARK_BLUE)
    row += 1

    card_bg = _pct_bg(overall_pct)
    card_fc = _pct_fg(overall_pct)
    for r in range(row, row + 4):
        for ci in range(2, 9):
            ws.cell(row=r, column=ci).fill   = _fill(card_bg)
            ws.cell(row=r, column=ci).border = _border()
    ws.merge_cells(start_row=row,   start_column=2, end_row=row,   end_column=8)
    ws.cell(row=row, column=2, value=study_label).font = Font(
        name='Arial', bold=True, size=10, color=C_DARK_BLUE)
    ws.cell(row=row, column=2).alignment = _center()
    ws.merge_cells(start_row=row+1, start_column=2, end_row=row+2, end_column=8)
    c = ws.cell(row=row+1, column=2, value=f"{overall_pct:.0f}%")
    c.font = Font(name='Arial', bold=True, size=32, color=card_fc)
    c.alignment = _center()
    ws.merge_cells(start_row=row+3, start_column=2, end_row=row+3, end_column=8)
    c = ws.cell(row=row+3, column=2,
                value=f"{t_total} / {g_total} objects vs. golden reference")
    c.font = Font(name='Arial', size=9, color=C_DARK_GREY)
    c.alignment = _center()
    for r in range(row, row + 4): ws.row_dimensions[r].height = 22
    row += 5

    # Compliance metrics table
    ws.merge_cells(f'B{row}:H{row}')
    ws.cell(row=row, column=2, value="COMPLIANCE METRICS").font = Font(
        name='Arial', bold=True, size=11, color=C_DARK_BLUE)
    row += 1

    _hdr_row(ws, row, ['Metric', 'Test Value', 'Golden Value', 'Status', 'Target', 'Gap', '', ''])
    row += 1

    def _metric_row(ws, r, metric, t_val, g_val, target="100%"):
        try:
            pv = float(str(t_val).replace('%','').replace('~','').strip())
        except Exception:
            pv = None
        try:
            gap_n = 100 - pv if pv is not None else None
            gap   = f"-{gap_n:.0f}pp" if gap_n is not None else '-'
        except Exception:
            gap = '-'
        status = ('OK' if (pv or 0) >= 90 else
                  'MEDIUM' if (pv or 0) >= 50 else
                  'LOW' if (pv or 0) >= 10 else 'CRITICAL')
        sf = {'CRITICAL': C_LIGHT_RED, 'LOW': C_LIGHT_AMB,
              'MEDIUM': "FFF2CC",       'OK':  C_LIGHT_GREEN}[status]
        vals = [metric, t_val, g_val, status, target, gap, '', '']
        for ci, v in enumerate(vals, 2):
            c = ws.cell(row=r, column=ci, value=v)
            c.font      = Font(name='Arial', size=9, bold=(ci == 5))
            c.fill      = _fill(sf if ci == 5 else (C_GREY if ci % 2 == 0 else C_WHITE))
            c.alignment = _left() if ci == 2 else _center(wrap=False)
            c.border    = _border()
        ws.row_dimensions[r].height = 18

    _metric_row(ws, row, "Overall content completeness",
                f"{overall_pct:.0f}%",   f"{100:.0f}% (golden is reference)")
    row += 1
    _metric_row(ws, row, "instanceType field coverage",
                f"{t_it_pct:.0f}%",      f"{g_it_pct:.0f}%")
    row += 1
    _metric_row(ws, row, "extensionAttributes presence (optional, best practice)",
                f"{t_ea_pct:.0f}%",      f"{g_ea_pct:.0f}%",  target="Optional (default=[])")
    row += 1
    _metric_row(ws, row, "USDM sequential ID compliance",
                f"{t_id_pct:.0f}%",      f"{g_id_pct:.0f}%")
    row += 1
    _metric_row(ws, row, "CDISC code system compliance",
                f"{t_cdisc_pct:.0f}%",   f"{g_cdisc_pct:.0f}%")
    row += 1
    _metric_row(ws, row, "USDM envelope complete (4 required keys)",
                "✓ Yes" if t_envelope_ok else "✗ No",
                "✓ Yes" if g_envelope_ok else "✗ No",
                target="✓ Yes")
    row += 2

    # Key findings
    ws.merge_cells(f'B{row}:H{row}')
    ws.cell(row=row, column=2, value="KEY FINDINGS").font = Font(
        name='Arial', bold=True, size=11, color=C_DARK_BLUE)
    row += 1

    _hdr_row(ws, row, ['Severity', 'Finding', 'Detail', '', '', '', '', ''],
             bg=C_MED_BLUE)
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=8)
    row += 1

    sev_colors = {
        'CRITICAL': (C_RED,    C_LIGHT_RED),
        'HIGH':     (C_ORANGE, C_LIGHT_RED),
        'MEDIUM':   ("7F6000", C_LIGHT_AMB),
        'LOW':      (C_GREEN,  C_LIGHT_GREEN),
    }

    def _finding_row(ws, r, sev, title, detail):
        fc, bg_c = sev_colors.get(sev, (C_DARK_GREY, C_GREY))
        c = ws.cell(row=r, column=2, value=sev)
        c.font = Font(name='Arial', bold=True, size=9, color=C_WHITE)
        c.fill = _fill(fc); c.alignment = _center(wrap=True); c.border = _border()
        c = ws.cell(row=r, column=3, value=title)
        c.font = Font(name='Arial', bold=True, size=9, color=C_DARK_BLUE)
        c.fill = _fill(bg_c); c.alignment = _left(wrap=True); c.border = _border()
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=8)
        c = ws.cell(row=r, column=4, value=detail)
        c.font = Font(name='Arial', size=9); c.fill = _fill(C_GREY)
        c.alignment = _left(wrap=True); c.border = _border()
        ws.row_dimensions[r].height = 42

    # Auto-generate findings from computed metrics
    if not t_envelope_ok:
        missing_env = REQUIRED_USDM_ENVELOPE_KEYS - set(t_root.keys())
        lower_note  = " ('Study' found — casing mismatch)" if not t_has_study_lower else ""
        _finding_row(ws, row, 'CRITICAL', 'Wrong USDM top-level envelope',
            f"Missing required keys: {', '.join(sorted(missing_env))}.{lower_note} "
            f"Also found {len(t_extra_keys)} extra non-USDM top-level key(s).")
        row += 1

    if t_extra_keys:
        _finding_row(ws, row, 'CRITICAL', f"{len(t_extra_keys)} extra non-USDM top-level section(s)",
            f"Keys present in test but not in USDM 4.0 spec: "
            f"{', '.join(t_extra_keys[:12])}{'...' if len(t_extra_keys) > 12 else ''}. "
            "These are likely internal pipeline artefacts and must be stripped before output.")
        row += 1

    if t_it_pct < 90:
        sev = 'CRITICAL' if t_it_pct < 10 else ('HIGH' if t_it_pct < 50 else 'MEDIUM')
        _finding_row(ws, row, sev,
            f"instanceType coverage: {t_it_pct:.0f}% (golden: {g_it_pct:.0f}%)",
            "USDM 4.0 requires 'instanceType' on every object. "
            f"Test has {t_it_has}/{t_it_tot} objects typed. "
            "Un-typed objects cannot be correctly parsed by USDM consumers.")
        row += 1

    if t_id_pct < 50:
        sev = 'CRITICAL' if t_id_pct == 0 else 'HIGH'
        _finding_row(ws, row, sev,
            f"ID format non-compliant: {t_id_pct:.0f}% USDM sequential (golden: {g_id_pct:.0f}%)",
            f"Test has {t_usdm_ids} USDM-style IDs and {t_hash_ids} hash-based IDs out of {t_tot_ids} total. "
            "USDM expects sequential type-prefixed IDs (e.g. 'Objective_1'). "
            "Hash-based IDs break referential integrity and human readability.")
        row += 1

    if t_cdisc_pct < 80:
        sev = 'CRITICAL' if t_cdisc_pct < 10 else 'HIGH'
        _finding_row(ws, row, sev,
            f"CDISC code system compliance: {t_cdisc_pct:.0f}% (golden: {g_cdisc_pct:.0f}%)",
            f"Test has {t_cdisc_c} CDISC codes, {t_nci_c} NCI codes, {t_other_c} other. "
            "USDM 4.0 requires codeSystem='http://www.cdisc.org' for standard codes. "
            "Replace NCI Thesaurus and other systems with CDISC CT equivalents.")
        row += 1

    if overall_pct < 90:
        sev = 'HIGH' if overall_pct >= 30 else 'CRITICAL'
        _finding_row(ws, row, sev,
            f"Content completeness: {overall_pct:.0f}% ({t_total} / {g_total} objects)",
            "Significant sections under-populated. See 'Content Completeness' sheet for per-section breakdown.")
        row += 1

    if t_ea_pct < g_ea_pct * 0.5:
        _finding_row(ws, row, 'LOW',
            f"extensionAttributes present on {t_ea_pct:.0f}% of objects (golden: {g_ea_pct:.0f}%)",
            "extensionAttributes is optional in the USDM 4.0 schema (default=[]) but is "
            "conventionally emitted on every object by compliant USDM serialisers. "
            "Consider adding it for full interoperability.")
        row += 1

    # ═════════════════════════════════════════════════════════════════════════
    # SHEET 2  Content Completeness
    # ═════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Content Completeness")
    ws2.sheet_view.showGridLines = False
    for ci, w in enumerate([30, 12, 12, 10, 14], 1): _cw(ws2, ci, w)

    row = _title_block(ws2, 1, "Content Completeness — Object Count vs. Golden", 5,
                       sub=f"Test JSON vs. Golden Reference  |  {study_label}")
    row += 1

    _hdr_row(ws2, row, ['Section / USDM Field', 'Golden', 'Test', 'Delta', '% Complete'])
    row += 1
    row = _sect_hdr(ws2, row, "◆  StudyDesign-level Collections", 5)

    def _data_row(ws, r, label, g_c, t_c, indent=0):
        pad   = "  " * indent
        delta = t_c - g_c
        pv    = _pct_val(t_c, g_c)
        ps    = _pct_str(t_c, g_c)
        for ci, v in enumerate([pad + label, g_c, t_c, delta, ps], 1):
            c = ws.cell(row=r, column=ci, value=v)
            c.border = _border()
            c.font   = Font(name='Arial', size=9)
            c.alignment = _left(wrap=False) if ci == 1 else _center()
            if ci == 4:
                c.fill = _fill(C_LIGHT_RED if delta < 0 else
                               (C_LIGHT_GREEN if delta > 0 else C_GREY))
            elif ci == 5:
                c.fill = _fill(_pct_bg(pv))
                c.font = Font(name='Arial', size=9, bold=True, color=_pct_fg(pv))
            else:
                c.fill = _fill(C_GREY if r % 2 == 0 else C_WHITE)
        ws.row_dimensions[r].height = 16

    for label, key, _ in SD_SECTIONS:
        _data_row(ws2, row, label, _cnt(g_sd, key), _cnt(t_sd, key), indent=1)
        row += 1

    row = _sect_hdr(ws2, row, "◆  StudyVersion-level Collections", 5)
    for label, key, _ in VER_SECTIONS:
        _data_row(ws2, row, label, _cnt(g_ver, key), _cnt(t_ver, key), indent=1)
        row += 1

    # Totals row
    row = _sect_hdr(ws2, row, "TOTAL", 5)
    for ci, v in enumerate(["TOTAL", g_total, t_total,
                             t_total - g_total, f"{overall_pct:.0f}%"], 1):
        c = ws2.cell(row=row, column=ci, value=v)
        c.font      = Font(name='Arial', bold=True, size=10)
        c.fill      = _fill(_pct_bg(overall_pct) if ci == 5 else C_LIGHT_BLUE)
        c.alignment = _left() if ci == 1 else _center()
        c.border    = _border()
    ws2.row_dimensions[row].height = 18

    # ═════════════════════════════════════════════════════════════════════════
    # SHEET 3  Schema Compliance
    # ═════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Schema Compliance")
    ws3.sheet_view.showGridLines = False
    for ci, w in enumerate([30, 18, 16, 14, 14, 30], 1): _cw(ws3, ci, w)

    row = _title_block(ws3, 1, "CDISC USDM 4.0 Schema Compliance Analysis", 6,
                       sub=f"{study_label}")
    row += 1

    # 1. Envelope
    row = _sect_hdr(ws3, row, "1. Top-Level USDM Envelope (4 required keys)", 6)
    _hdr_row(ws3, row, ['Required Key', 'In Golden?', 'In Test?', 'Compliant?', 'Notes', ''])
    row += 1

    envelope_rows = [
        ('study',         g_key == 'study',            t_key == 'study' or 'study' in t_root,
         f"Test key is '{t_key}'" if t_key != 'study' else "Correct"),
        ('usdmVersion',   'usdmVersion' in g_root,     'usdmVersion' in t_root,    ""),
        ('systemName',    'systemName' in g_root,      'systemName' in t_root,     ""),
        ('systemVersion', 'systemVersion' in g_root,   'systemVersion' in t_root,  ""),
    ]
    for key, in_g, in_t, note in envelope_rows:
        for ci, v in enumerate([key,
                                 '✓' if in_g else '✗',
                                 '✓' if in_t else '✗',
                                 '✓' if in_t else '✗ Non-compliant',
                                 note, ''], 1):
            c = ws3.cell(row=row, column=ci, value=v)
            c.border = _border(); c.font = Font(name='Arial', size=9)
            c.alignment = _left() if ci in (1, 5) else _center()
            if ci == 3: c.fill = _fill(C_LIGHT_GREEN if in_g else C_LIGHT_RED)
            elif ci in (4, 5):
                c.fill = _fill(C_LIGHT_GREEN if in_t else C_LIGHT_RED)
                c.font = Font(name='Arial', size=9, bold=True,
                              color=C_GREEN if in_t else C_RED)
        ws3.row_dimensions[row].height = 18
        row += 1

    # Extra top-level keys
    if t_extra_keys:
        row += 1
        row = _sect_hdr(ws3, row, f"1b. Extra Non-USDM Top-Level Keys in Test ({len(t_extra_keys)} found)", 6)
        _hdr_row(ws3, row, ['Key Name', 'Type', '', 'Disposition', '', ''])
        row += 1
        for k in t_extra_keys:
            v_type = type(t_root[k]).__name__
            for ci, val in enumerate([k, v_type, '', 'Strip — not part of USDM 4.0 envelope', '', ''], 1):
                c = ws3.cell(row=row, column=ci, value=val)
                c.border = _border(); c.font = Font(name='Arial', size=9)
                c.fill   = _fill(C_LIGHT_AMB); c.alignment = _left()
            ws3.row_dimensions[row].height = 16
            row += 1

    # 2. Object-level mandatory fields
    row += 1
    row = _sect_hdr(ws3, row, "2. Object-Level Fields (USDM 4.0 compliance)", 6)
    _hdr_row(ws3, row, ['Field', 'Scope', 'Golden', 'Test', 'Impact', ''])
    row += 1

    obj_field_rows = [
        ('instanceType',
         'Every USDM object',
         f"{g_it_pct:.0f}%",
         f"{t_it_pct:.0f}%",
         'Objects cannot be typed — breaks model parsing' if t_it_pct < 90 else 'OK'),
        ('id (sequential TypeName_N format)',
         'Every USDM object',
         f"{g_id_pct:.0f}%",
         f"{t_id_pct:.0f}%",
         'Referential integrity broken; IDs not human-readable' if t_id_pct < 90 else 'OK'),
        ('codeSystem = http://www.cdisc.org',
         'All Code objects',
         f"{g_cdisc_pct:.0f}%",
         f"{t_cdisc_pct:.0f}%",
         'Non-compliant terminology; NCI/mixed systems used' if t_cdisc_pct < 90 else 'OK'),
        ('extensionAttributes (optional, default=[])',
         'Every USDM object',
         f"{g_ea_pct:.0f}%",
         f"{t_ea_pct:.0f}%",
         'Optional per schema; conventionally emitted by compliant serialisers'),
    ]
    for fname, scope, g_comp, t_comp, impact in obj_field_rows:
        for ci, v in enumerate([fname, scope, g_comp, t_comp, impact, ''], 1):
            c = ws3.cell(row=row, column=ci, value=v)
            c.border = _border(); c.font = Font(name='Arial', size=9)
            c.alignment = _left(wrap=True) if ci in (1, 2, 5) else _center(wrap=True)
            if ci == 4:
                try: pv = float(str(v).replace('%','').strip())
                except Exception: pv = None
                c.fill = _fill(_pct_bg(pv))
                c.font = Font(name='Arial', size=9, bold=True, color=_pct_fg(pv))
        ws3.row_dimensions[row].height = 28
        row += 1

    # ═════════════════════════════════════════════════════════════════════════
    # SHEET 4  Field-Level Diff
    # ═════════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Field-Level Diff")
    ws4.sheet_view.showGridLines = False
    for ci, w in enumerate([28, 16, 10, 10, 36], 1): _cw(ws4, ci, w)

    row = _title_block(ws4, 1, "Field-Level Structural Diff — USDM Object Keys", 5,
                       sub=f"{study_label}")
    row += 1

    def _diff_table(ws, r, title, g_keys, t_keys):
        r = _sect_hdr(ws, r, title, 5)
        _hdr_row(ws, r, ['Field', 'Status', 'In Golden?', 'In Test?', 'Notes'])
        r += 1
        for field in sorted(g_keys | t_keys):
            in_g = field in g_keys
            in_t = field in t_keys
            if in_g and in_t:
                status, bg, note = "✓ Match",    C_LIGHT_GREEN, ""
            elif in_g and not in_t:
                status, bg, note = "⚠ Missing",  C_LIGHT_RED,   "Required by golden — add to test"
            else:
                status, bg, note = "➕ Extra",    C_LIGHT_AMB,   "Not in golden — non-USDM field"
            for ci, v in enumerate([field, status,
                                     '✓' if in_g else '',
                                     '✓' if in_t else '',
                                     note], 1):
                c = ws.cell(row=r, column=ci, value=v)
                c.border = _border(); c.font = Font(name='Arial', size=9)
                c.fill   = _fill(bg)
                c.alignment = _center() if ci in (2, 3, 4) else _left()
            ws.row_dimensions[r].height = 16
            r += 1
        return r + 1

    row = _diff_table(ws4, row, "A. Top-Level Envelope",
                      g_top_keys, REQUIRED_USDM_ENVELOPE_KEYS)
    row = _diff_table(ws4, row, "B. StudyVersion Object Fields", g_ver_keys, t_ver_keys)
    row = _diff_table(ws4, row, "C. StudyDesign Object Fields",  g_sd_keys,  t_sd_keys)

    # ═════════════════════════════════════════════════════════════════════════
    # SHEET 5  Content Quality
    # ═════════════════════════════════════════════════════════════════════════
    ws5 = wb.create_sheet("Content Quality")
    ws5.sheet_view.showGridLines = False
    for ci, w in enumerate([28, 44, 44], 1): _cw(ws5, ci, w)

    row = _title_block(ws5, 1, "Content Quality — Key Field Spot-Check", 3,
                       sub=f"{study_label}")
    row += 1

    def _cq_hdr(ws, r, cols):
        _hdr_row(ws, r, cols); ws.row_dimensions[r].height = 20; return r + 1

    def _cq_row(ws, r, field, g_val, t_val):
        gv = str(g_val)[:140] if g_val is not None else 'None'
        tv = str(t_val)[:140] if t_val is not None else 'None'
        # Simple match: first 10 chars match (enough to catch same vs. different)
        match = str(g_val)[:10] == str(t_val)[:10] if g_val and t_val else (g_val == t_val)
        for ci, (v, is_val) in enumerate([(field, False), (gv, True), (tv, True)], 1):
            c = ws5.cell(row=r, column=ci, value=v)
            c.border = _border(); c.font = Font(name='Arial', size=9)
            c.alignment = _left(wrap=True)
            if is_val:
                c.fill = _fill(C_LIGHT_GREEN if match else
                               (C_LIGHT_RED if ci == 3 else C_LIGHT_BLUE))
            else:
                c.fill = _fill(C_GREY)
        ws5.row_dimensions[r].height = 32

    # Metadata
    row = _sect_hdr(ws5, row, "Study Metadata", 3)
    row = _cq_hdr(ws5, row, ['Field', 'Golden Value', 'Test Value'])
    for field, gv, tv in [
        ('study.name',       g_s.get('name'),         t_s.get('name')),
        ('study.label',      g_s.get('label'),        t_s.get('label')),
        ('versionIdentifier',g_ver.get('versionIdentifier'),
                             t_ver.get('versionIdentifier')),
        ('usdmVersion',      g_root.get('usdmVersion'), t_root.get('usdmVersion')),
        ('systemName',       g_root.get('systemName'),  t_root.get('systemName')),
        ('Titles count',     len(g_ver.get('titles',[])),
                             len(t_ver.get('titles',[]))),
        ('Brief Study Title',get_brief_title(g_ver),   get_brief_title(t_ver)),
    ]:
        _cq_row(ws5, row, field, gv, tv); row += 1

    row += 1
    row = _sect_hdr(ws5, row, "Study Design Summary — Object Counts", 3)
    row = _cq_hdr(ws5, row, ['Field', 'Golden Count', 'Test Count'])
    g_pop = g_sd.get('population', {}); t_pop = t_sd.get('population', {})
    for field, gv, tv in [
        ('studyType',
         g_sd.get('studyType', {}).get('decode', '') if isinstance(g_sd.get('studyType'), dict) else g_sd.get('studyType', ''),
         t_sd.get('studyType', {}).get('decode', '') if isinstance(t_sd.get('studyType'), dict) else t_sd.get('studyType', '')),
        ('population.label',          g_pop.get('label',''),  t_pop.get('label','')),
        ('plannedEnrollmentNumber',
         g_pop.get('plannedEnrollmentNumber', {}).get('value', '')
             if isinstance(g_pop.get('plannedEnrollmentNumber'), dict)
             else g_pop.get('plannedEnrollmentNumber', ''),
         t_pop.get('plannedEnrollmentNumber', {}).get('value', '')
             if isinstance(t_pop.get('plannedEnrollmentNumber'), dict)
             else t_pop.get('plannedEnrollmentNumber', 'MISSING')),
        ('Arms',              len(g_sd.get('arms',[])),            len(t_sd.get('arms',[]))),
        ('Epochs',            len(g_sd.get('epochs',[])),          len(t_sd.get('epochs',[]))),
        ('Encounters',        len(g_sd.get('encounters',[])),      len(t_sd.get('encounters',[]))),
        ('Activities',        len(g_sd.get('activities',[])),      len(t_sd.get('activities',[]))),
        ('Objectives',        len(g_sd.get('objectives',[])),      len(t_sd.get('objectives',[]))),
        ('EC items',          len(g_ver.get('eligibilityCriterionItems',[])),
                              len(t_ver.get('eligibilityCriterionItems',[]))),
        ('Schedule Timelines',len(g_sd.get('scheduleTimelines',[])),
                              len(t_sd.get('scheduleTimelines',[]))),
        ('Biomedical Concepts',len(g_ver.get('biomedicalConcepts',[])),
                              len(t_ver.get('biomedicalConcepts',[]))),
    ]:
        _cq_row(ws5, row, field, gv, tv); row += 1

    row += 1
    def _objective_text(obj):
        """Return the text value from studyDesigns -> objectives -> text,
        cleaned of any HTML/USDM template tags. Returns 'N/A' if absent."""
        if not obj:
            return 'N/A'
        raw = (obj.get('text') or '').strip()
        # Strip HTML tags (e.g. <p>, <br>) and USDM template tags
        raw = re.sub(r'<[^>]+>', ' ', raw)
        raw = ' '.join(raw.split())   # collapse whitespace
        return raw[:250] if raw else 'N/A'

    row = _sect_hdr(ws5, row, "Objectives (first 5 — name and text)", 3)
    row = _cq_hdr(ws5, row, ['Field', 'Golden Value', 'Test Value'])
    g_objs = g_sd.get('objectives', [])
    t_objs = t_sd.get('objectives', [])
    n_objs = min(5, max(len(g_objs), len(t_objs), 1))
    for i in range(n_objs):
        go = g_objs[i] if i < len(g_objs) else None
        to = t_objs[i] if i < len(t_objs) else None
        _cq_row(ws5, row, f"Objective {i+1} — name",
                go.get('name', '') if go else 'N/A',
                to.get('name', '') if to else 'MISSING')
        row += 1
        _cq_row(ws5, row, f"Objective {i+1} — text",
                _objective_text(go),
                _objective_text(to) if to else 'MISSING')
        row += 1

    row += 1
    row = _sect_hdr(ws5, row, "Eligibility Criteria (first 3 IC, first 3 EC from golden)", 3)
    row = _cq_hdr(ws5, row, ['Field', 'Golden Value', 'Test Value'])

    # Category is on EligibilityCriterion objects (studyDesign.eligibilityCriteria),
    # not on EligibilityCriterionItem objects (studyVersion.eligibilityCriterionItems).
    # We look up the criterion text via criterionItemId.
    def _build_ec_lookup(sd_obj, ver_obj):
        """
        Returns (inclusion_list, exclusion_list) of criterion text strings.
        Four-level resolution chain per criterion:
          1. ec_obj.text  (direct text field on EligibilityCriterion)
          2. criterionItemId -> EligibilityCriterionItem.text  (exact ID match)
          3. Positional match against uncategorised items pool (handles broken ID links)
          4. ec_obj.name  (short name, last resort)
        """
        all_items = [item for item in ver_obj.get('eligibilityCriterionItems', [])
                     if isinstance(item, dict)]
        items_by_id = {item['id']: item for item in all_items if 'id' in item}

        def _clean(text):
            text = re.sub(r'<[^>]+>', ' ', text or '')
            text = re.sub(r'\s+', ' ', text).strip()
            return text[:250] if text else ''

        # Build separate categorised pools for positional fallback
        ic_pool = [it for it in all_items
                   if isinstance(it.get('category'), dict)
                   and it['category'].get('code') == 'C25532']
        ec_pool = [it for it in all_items
                   if isinstance(it.get('category'), dict)
                   and it['category'].get('code') == 'C25370']
        # Items with no category (common in test JSONs) used positionally
        uncategorised = [it for it in all_items
                         if not isinstance(it.get('category'), dict)]

        counters = {'ic': 0, 'ec': 0, 'unc': 0}

        def _resolve(ec_obj, is_ic):
            # 1. Direct text
            text = _clean(ec_obj.get('text', ''))
            if text:
                return text
            # 2. Exact criterionItemId link
            item_id = ec_obj.get('criterionItemId')
            if item_id and item_id in items_by_id:
                text = _clean(items_by_id[item_id].get('text', ''))
                if text:
                    return text
            # 3. Positional match against uncategorised items (test JSONs often
            #    have items with no category, stored IC-first then EC-first)
            if uncategorised:
                idx = counters['unc']
                if idx < len(uncategorised):
                    text = _clean(uncategorised[idx].get('text', ''))
                    counters['unc'] += 1
                    if text:
                        return text
            # 4. Positional match against categorised pool
            pool = ic_pool if is_ic else ec_pool
            key  = 'ic' if is_ic else 'ec'
            idx  = counters[key]
            if idx < len(pool):
                text = _clean(pool[idx].get('text', ''))
                counters[key] += 1
                if text:
                    return text
            # 5. Name as last resort
            return _clean(ec_obj.get('name', '')) or 'N/A'

        ic, ec = [], []
        for criterion in sd_obj.get('eligibilityCriteria', []):
            if not isinstance(criterion, dict):
                continue
            cat  = criterion.get('category')
            code = cat.get('code', '') if isinstance(cat, dict) else ''
            if code == 'C25532':
                ic.append(_resolve(criterion, is_ic=True))
            elif code == 'C25370':
                ec.append(_resolve(criterion, is_ic=False))
        return ic, ec

    g_ic, g_ec = _build_ec_lookup(g_sd, g_ver)
    t_ic, t_ec = _build_ec_lookup(t_sd, t_ver)

    for i in range(3):
        gv = g_ic[i] if i < len(g_ic) else 'N/A'
        tv = t_ic[i] if i < len(t_ic) else 'MISSING'
        _cq_row(ws5, row, f"Inclusion Criterion {i+1}", gv, tv)
        row += 1
    for i in range(3):
        gv = g_ec[i] if i < len(g_ec) else 'N/A'
        tv = t_ec[i] if i < len(t_ec) else 'MISSING'
        _cq_row(ws5, row, f"Exclusion Criterion {i+1}", gv, tv)
        row += 1

    # ═════════════════════════════════════════════════════════════════════════
    # SHEET 6  Recommendations
    # ═════════════════════════════════════════════════════════════════════════
    ws6 = wb.create_sheet("Recommendations")
    ws6.sheet_view.showGridLines = False
    for ci, w in enumerate([6, 34, 18, 10, 14, 52], 1): _cw(ws6, ci, w)

    row = _title_block(ws6, 1, "Remediation Recommendations — Priority Order", 6,
                       sub=f"{study_label}")
    row += 1

    _hdr_row(ws6, row,
             ['#', 'Issue', 'Category', 'Effort', 'Priority', 'Recommended Action'])
    row += 1

    pri_fill = {'P1 - Critical': C_RED, 'P2 - High': C_ORANGE, 'P3 - Medium': C_AMBER}
    eff_fill = {'Low': C_LIGHT_GREEN, 'Medium': C_LIGHT_AMB, 'High': C_LIGHT_RED}

    # Auto-generate recommendations based on computed metrics
    recs = []

    if not t_envelope_ok:
        recs.append((
            "Fix USDM top-level envelope", "Schema Structure", "Low", "P1 - Critical",
            f"Rename '{t_key}' → 'study'. "
            "Add 'usdmVersion': '4.0.0', 'systemName', and 'systemVersion' to the root JSON. "
            "Strip all non-USDM top-level keys from the output."
        ))

    if t_it_pct < 90:
        recs.append((
            f"Add instanceType to all objects (currently {t_it_pct:.0f}%)",
            "Schema Compliance", "Medium", "P1 - Critical",
            "Every dict object in the USDM graph must carry 'instanceType' matching its class name "
            "(e.g. 'Study', 'StudyDesign', 'Objective', 'Code'). Implement in the serialisation layer."
        ))

    if t_id_pct < 90:
        recs.append((
            f"Reformat IDs to USDM sequential style (currently {t_id_pct:.0f}%)",
            "ID Compliance", "Medium", "P1 - Critical",
            "Replace hash-based IDs with sequential type-prefixed IDs (e.g. 'Objective_1'). "
            "Maintain a per-type counter in the serialiser to ensure uniqueness."
        ))

    if t_cdisc_pct < 80:
        recs.append((
            f"Standardise code system to CDISC (currently {t_cdisc_pct:.0f}%)",
            "Terminology", "High", "P1 - Critical",
            "Map all codes to CDISC CT: codeSystem='http://www.cdisc.org', "
            "codeSystemVersion='YYYY-MM-DD'. Replace NCI Thesaurus and other ad-hoc systems."
        ))

    # Per-section content completeness recommendations
    section_gaps = []
    for label, key, level in SD_SECTIONS:
        g_c = _cnt(g_sd, key); t_c = _cnt(t_sd, key)
        pv  = _pct_val(t_c, g_c)
        if pv is not None and pv < 50 and g_c > 0:
            section_gaps.append((label, g_c, t_c, pv))

    for label, key, level in VER_SECTIONS:
        g_c = _cnt(g_ver, key); t_c = _cnt(t_ver, key)
        pv  = _pct_val(t_c, g_c)
        if pv is not None and pv < 50 and g_c > 0:
            section_gaps.append((label, g_c, t_c, pv))

    # Sort by severity (lowest pct first)
    section_gaps.sort(key=lambda x: x[3])
    for label, g_c, t_c, pv in section_gaps[:8]:
        priority = "P1 - Critical" if pv < 10 else "P2 - High"
        recs.append((
            f"Populate {label} ({t_c}/{g_c} = {pv:.0f}%)",
            "Content Completeness", "High", priority,
            f"Golden has {g_c} {label.lower()} object(s); test has {t_c}. "
            "Extract all relevant data from the protocol and map to the correct USDM objects."
        ))

    if t_ea_pct < g_ea_pct * 0.5:
        recs.append((
            "Add extensionAttributes to all objects (best practice)",
            "Interoperability", "Low", "P3 - Medium",
            "extensionAttributes is optional in the schema (default=[]) but conventionally "
            "emitted on every object by compliant USDM serialisers. Add it for full interoperability."
        ))

    if t_extra_keys:
        recs.append((
            f"Remove {len(t_extra_keys)} non-USDM top-level section(s)",
            "Schema Compliance", "Low", "P3 - Medium",
            "Strip internal pipeline artefacts from the USDM output envelope: "
            f"{', '.join(t_extra_keys[:10])}. "
            "Optionally map relevant content to USDM narrativeContentItems or scheduleTimelines."
        ))

    for idx, (issue, cat, effort, priority, action) in enumerate(recs, 1):
        for ci, v in enumerate([idx, issue, cat, effort, priority, action], 1):
            c = ws6.cell(row=row, column=ci, value=v)
            c.border = _border(); c.font = Font(name='Arial', size=9)
            c.alignment = _center() if ci in (1, 4) else _left(wrap=True)
            if ci == 1:
                c.fill = _fill(C_LIGHT_BLUE)
                c.font = Font(name='Arial', bold=True, size=9)
            elif ci == 4:
                c.fill = _fill(eff_fill.get(effort, C_GREY))
            elif ci == 5:
                c.fill = _fill(pri_fill.get(priority, C_GREY))
                c.font = Font(name='Arial', bold=True, size=9, color=C_WHITE)
                c.alignment = _center()
        ws6.row_dimensions[row].height = 42
        row += 1

    # ── Save safely via temp file → validate → atomic move ───────────────────
    # Writing directly to a OneDrive/network folder can cause corruption if the
    # file is already open in Excel or a sync agent grabs it mid-write.
    # Strategy: write to a local temp file, verify it is a valid xlsx (zip),
    # then move it into place in one atomic operation.
    tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx",
                                        dir=tempfile.gettempdir(),
                                        prefix="usdm_report_")
    os.close(tmp_fd)
    try:
        wb.save(tmp_path)

        # Validate: a valid .xlsx is a ZIP archive containing [Content_Types].xml
        try:
            with zipfile.ZipFile(tmp_path, 'r') as z:
                if '[Content_Types].xml' not in z.namelist():
                    raise ValueError("Missing [Content_Types].xml — not a valid xlsx")
        except (zipfile.BadZipFile, ValueError) as exc:
            os.remove(tmp_path)
            sys.exit(f"ERROR: Generated file failed validation: {exc}")

        # Atomic move to final destination
        shutil.move(tmp_path, str(output_path))

    except Exception:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
        raise

    print(f"  Report saved → {output_path}")
    print(f"  Sheets: {wb.sheetnames}")

    # ── Print summary to console ──────────────────────────────────────────────
    print()
    print("  ┌─────────────────────────────────────────────────────────────┐")
    print(f"  │  Study : {study_label[:55]:<55} │")
    print(f"  │  Content completeness  : {overall_pct:>5.1f}%  ({t_total}/{g_total} objects)     │")
    print(f"  │  instanceType coverage : {t_it_pct:>5.1f}%  (golden: {g_it_pct:.0f}%)           │")
    print(f"  │  USDM sequential IDs   : {t_id_pct:>5.1f}%                              │")
    print(f"  │  CDISC code system     : {t_cdisc_pct:>5.1f}%                              │")
    print(f"  │  USDM envelope OK      : {'Yes' if t_envelope_ok else 'No — ' + str(len(REQUIRED_USDM_ENVELOPE_KEYS - set(t_root.keys()))) + ' key(s) missing':<38} │")
    print(f"  │  Extra top-level keys  : {len(t_extra_keys):<38} │")
    print(f"  │  Recommendations       : {len(recs):<38} │")
    print("  └─────────────────────────────────────────────────────────────┘")


# ──────────────────────────────────────────────────────────────────────────────
# Entry point
# ──────────────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description="Generate a CDISC USDM 4.0 deviation report comparing a test JSON "
                    "against a golden reference JSON.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python usdm_deviation_report.py golden.json test.json
  python usdm_deviation_report.py golden.json test.json my_report.xlsx
  python usdm_deviation_report.py ./data/CDISC_golden.json ./data/CDISC_test.json
        """)
    parser.add_argument("golden",  help="Path to the golden reference JSON file")
    parser.add_argument("test",    help="Path to the test extraction JSON file")
    parser.add_argument("output",  nargs="?",
                        help="Output .xlsx path (default: <golden_stem>_deviation_report.xlsx "
                             "in the same directory as the golden file)")
    args = parser.parse_args()

    golden_path = Path(args.golden)
    test_path   = Path(args.test)

    if not golden_path.exists():
        sys.exit(f"ERROR: Golden file not found: {golden_path}")
    if not test_path.exists():
        sys.exit(f"ERROR: Test file not found: {test_path}")

    if args.output:
        output_path = Path(args.output)
    else:
        output_path = golden_path.parent / (golden_path.stem + "_deviation_report.xlsx")

    # Always enforce .xlsx extension
    if output_path.suffix.lower() != ".xlsx":
        output_path = output_path.with_suffix(".xlsx")

    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Warn if the destination file is currently open in Excel
    lock_file = output_path.parent / ("~$" + output_path.name)
    if lock_file.exists():
        print(f"  WARNING: '{output_path.name}' appears to be open in Excel.")
        print("           Close it before running this script to avoid conflicts.")
        print()

    print(f"\nUSDM 4.0 Deviation Report Generator")
    print(f"  Golden : {golden_path}")
    print(f"  Test   : {test_path}")
    print(f"  Output : {output_path}")
    print()

    build_report(golden_path, test_path, output_path)
    print("\nDone.\n")


if __name__ == "__main__":
    main()
